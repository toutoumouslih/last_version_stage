from rest_framework import viewsets
from rest_framework import generics
from rest_framework.decorators import api_view
from rest_framework.response import Response
import openpyxl
from .models import Country, Region, Department, Commune, DemographicData,  EducationLevel, Census
from .serializers import CountrySerializer, RegionSerializer, DepartmentSerializer, CommuneSerializer ,  DemographicDataSerializer
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect
from django.contrib import messages
from import_export.formats import base_formats
from .resources import DemographicDataResource, EducationLevelResource
from tablib import Dataset
from django.contrib.auth.decorators import user_passes_test
from django.utils.decorators import method_decorator
from django.views.generic import FormView
from .forms import CentralizedImportForm
from django.urls import reverse_lazy
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from datetime import datetime


def normalize_header(header):
    return header.strip().lower().replace(' ', '_000_temp_replace_').replace('+', 'plus').replace('%', 'percent').replace('.', '_dot_').replace('-', '_dash_').replace('\'', '').replace('(', '').replace(')', '').replace('\\', '').replace('/', '').replace('É', 'e').replace('é', 'e').replace('à', 'a').replace('À', 'a').replace('â', 'a').replace('Â', 'a').replace('ê', 'e').replace('Ê', 'e').replace('î', 'i').replace('Î', 'i').replace('ô', 'o').replace('Ô', 'o').replace('ù', 'u').replace('Ù', 'u').replace('û', 'u').replace('Û', 'u').replace('ç', 'c').replace('Ç', 'c').replace('ë', 'e').replace('Ë', 'e').replace('ï', 'i').replace('Ï', 'i').replace('ü', 'u').replace('Ü', 'u').replace('ñ', 'n').replace('Ñ', 'n').replace('_000_temp_replace_', '')

def get_header_index(headers, expected_header_name):
    normalized_expected = normalize_header(expected_header_name)
    for i, header in enumerate(headers):
        if normalize_header(header) == normalized_expected:
            return i
    return -1 # Return -1 if not found

def staff_required(login_url=None):
    return user_passes_test(lambda u: u.is_staff, login_url=login_url)

@method_decorator(staff_member_required, name='dispatch')
class CentralizedImportView(FormView):
    template_name = 'admin/import_data.html'
    form_class = CentralizedImportForm
    success_url = reverse_lazy('admin:index')

    def form_valid(self, form):
        census = form.cleaned_data['census_year']
        selected_content_type = form.cleaned_data['file_format']
        
        # Message de début d'import
        messages.info(self.request, f"Début de l'importation des données pour le recensement {census.year}...")
        
        file_format_instance = None
        format_extension = None
        
        # Mappage des types de contenu aux extensions
        if selected_content_type == base_formats.CSV().CONTENT_TYPE:
            format_extension = 'csv'
            file_format_instance = base_formats.CSV()
        elif selected_content_type == base_formats.XLS().CONTENT_TYPE:
            format_extension = 'xls'
            file_format_instance = base_formats.XLS()
        elif selected_content_type == base_formats.XLSX().CONTENT_TYPE:
            format_extension = 'xlsx'
            file_format_instance = base_formats.XLSX()
        elif selected_content_type == base_formats.JSON().CONTENT_TYPE:
            format_extension = 'json'
            file_format_instance = base_formats.JSON()

        if not file_format_instance or not format_extension:
            messages.error(self.request, f"Format de fichier non valide : {selected_content_type}")
            return self.render_to_response(self.get_context_data(form=form))

        dataset = Dataset()
        imported_file = self.request.FILES['import_file']
        
        try:
            # Lecture du fichier
            if format_extension in ['csv', 'json']:
                file_content = imported_file.read().decode('utf-8')
            else:
                file_content = imported_file.read()

            data = dataset.load(file_content, format=format_extension)
            total_rows = len(data)
            messages.info(self.request, f"Fichier chargé avec succès : {total_rows} lignes trouvées")
            
            # Import des données démographiques
            demographic_resource = DemographicDataResource()
            demographic_data_rows = []
            skipped_rows = 0
            
            for row_index, row_data in enumerate(data):
                # Vérification du niveau_donnee
                niveau_idx = get_header_index(data.headers, 'niveau_donnee')
                niveau = row_data[niveau_idx].strip().lower() if niveau_idx != -1 and row_data[niveau_idx] else ''
                if niveau not in ['region', 'departement', 'commune']:
                    skipped_rows += 1
                    continue
                # Récupération des codes
                country_code = row_data[get_header_index(data.headers, 'Country Code')] if get_header_index(data.headers, 'Country Code') != -1 else None
                region_code = row_data[get_header_index(data.headers, 'Region Code')] if get_header_index(data.headers, 'Region Code') != -1 else None
                department_code = row_data[get_header_index(data.headers, 'Department Code')] if get_header_index(data.headers, 'Department Code') != -1 else None
                commune_code = row_data[get_header_index(data.headers, 'Commune Code')] if get_header_index(data.headers, 'Commune Code') != -1 else None

                # Vérification des codes
                if not all([country_code, region_code]):
                    skipped_rows += 1
                    continue

                # Récupération des instances
                country_instance = Country.objects.filter(code=country_code).first()
                region_instance = Region.objects.filter(adm1_pcode=region_code).first()
                department_instance = Department.objects.filter(adm2_pcode=department_code).first() if department_code else None
                commune_instance = Commune.objects.filter(adm3_pcode=commune_code).first() if commune_code else None

                if not country_instance or not region_instance:
                    skipped_rows += 1
                    continue

                # Construction des données démographiques avec validation des valeurs
                try:
                    demographic_row = {
                        'census': census.pk,
                        'country': country_instance.pk,
                        'region': region_instance.pk,
                        'department': department_instance.pk if department_instance else None,
                        'commune': commune_instance.pk if commune_instance else None,
                        'total_population': int(row_data[get_header_index(data.headers, 'Total Population')]) if get_header_index(data.headers, 'Total Population') != -1 else None,
                        'male_percentage': min(float(row_data[get_header_index(data.headers, 'Male Population')]), 100) if get_header_index(data.headers, 'Male Population') != -1 else None,
                        'female_percentage': min(float(row_data[get_header_index(data.headers, 'Female Population')]), 100) if get_header_index(data.headers, 'Female Population') != -1 else None,
                        'urban_percentage': min(float(row_data[get_header_index(data.headers, 'Urban Population')]), 100) if get_header_index(data.headers, 'Urban Population') != -1 else 0,
                        'rural_percentage': min(float(row_data[get_header_index(data.headers, 'Rural Population')]), 100) if get_header_index(data.headers, 'Rural Population') != -1 else 0,
                        'population_10_plus': int(row_data[get_header_index(data.headers, 'Population 10+')]) if get_header_index(data.headers, 'Population 10+') != -1 else None,
                        'single_rate': min(float(row_data[get_header_index(data.headers, 'Single Rate')]), 100) if get_header_index(data.headers, 'Single Rate') != -1 else None,
                        'married_rate': min(float(row_data[get_header_index(data.headers, 'Married Rate')]), 100) if get_header_index(data.headers, 'Married Rate') != -1 else None,
                        'divorced_rate': min(float(row_data[get_header_index(data.headers, 'Divorced Rate')]), 100) if get_header_index(data.headers, 'Divorced Rate') != -1 else None,
                        'widowed_rate': min(float(row_data[get_header_index(data.headers, 'Widowed Rate')]), 100) if get_header_index(data.headers, 'Widowed Rate') != -1 else None,
                        'school_enrollment_rate': min(float(row_data[get_header_index(data.headers, 'School Enrollment Rate')]), 100) if get_header_index(data.headers, 'School Enrollment Rate') != -1 else None,
                        'illiteracy_rate_10_plus': min(float(row_data[get_header_index(data.headers, 'Illiteracy Rate 10+')]), 100) if get_header_index(data.headers, 'Illiteracy Rate 10+') != -1 else None,
                        'population_15_plus': int(row_data[get_header_index(data.headers, 'Population 15+')]) if get_header_index(data.headers, 'Population 15+') != -1 else None,
                        'illiteracy_rate_15_plus': min(float(row_data[get_header_index(data.headers, 'Illiteracy Rate 15+')]), 100) if get_header_index(data.headers, 'Illiteracy Rate 15+') != -1 else None,
                    }
                    demographic_data_rows.append(demographic_row)
                except (ValueError, TypeError):
                    skipped_rows += 1
                    continue

            # Import des données démographiques
            if demographic_data_rows:
                demographic_dataset = Dataset()
                demographic_dataset.dict = demographic_data_rows
                demographic_result = demographic_resource.import_data(demographic_dataset, dry_run=False, raise_errors=True)
                
                # Import des données d'éducation
                education_resource = EducationLevelResource()
                education_data_rows = []
                education_imported = 0
                education_skipped = 0

                for row_index, row_data in enumerate(data):
                    # Recherche de l'instance démographique correspondante
                    demographic_instance = DemographicData.objects.filter(
                        census=census,
                        country__code=row_data[get_header_index(data.headers, 'Country Code')] if get_header_index(data.headers, 'Country Code') != -1 else None,
                        region__adm1_pcode=row_data[get_header_index(data.headers, 'Region Code')] if get_header_index(data.headers, 'Region Code') != -1 else None,
                        department__adm2_pcode=row_data[get_header_index(data.headers, 'Department Code')] if get_header_index(data.headers, 'Department Code') != -1 else None,
                        commune__adm3_pcode=row_data[get_header_index(data.headers, 'Commune Code')] if get_header_index(data.headers, 'Commune Code') != -1 else None,
                    ).first()

                    if demographic_instance:
                        try:
                            # Vérifier si des données d'éducation existent déjà
                            existing_education = EducationLevel.objects.filter(demographic_data=demographic_instance).first()
                            if existing_education:
                                # Mettre à jour les données existantes
                                education_row = {
                                    'id': existing_education.id,
                                    'demographic_data': demographic_instance.pk,
                                    'no_education': min(float(row_data[get_header_index(data.headers, 'No Education')]), 100) if get_header_index(data.headers, 'No Education') != -1 else None,
                                    'preschool': min(float(row_data[get_header_index(data.headers, 'Preschool')]), 100) if get_header_index(data.headers, 'Preschool') != -1 else None,
                                    'primary': min(float(row_data[get_header_index(data.headers, 'Primary')]), 100) if get_header_index(data.headers, 'Primary') != -1 else None,
                                    'middle_school': min(float(row_data[get_header_index(data.headers, 'Middle School')]), 100) if get_header_index(data.headers, 'Middle School') != -1 else None,
                                    'high_school': min(float(row_data[get_header_index(data.headers, 'High School')]), 100) if get_header_index(data.headers, 'High School') != -1 else None,
                                    'university': min(float(row_data[get_header_index(data.headers, 'University')]), 100) if get_header_index(data.headers, 'University') != -1 else None,
                                }
                            else:
                                # Créer de nouvelles données
                                education_row = {
                                    'demographic_data': demographic_instance.pk,
                                    'no_education': min(float(row_data[get_header_index(data.headers, 'No Education')]), 100) if get_header_index(data.headers, 'No Education') != -1 else None,
                                    'preschool': min(float(row_data[get_header_index(data.headers, 'Preschool')]), 100) if get_header_index(data.headers, 'Preschool') != -1 else None,
                                    'primary': min(float(row_data[get_header_index(data.headers, 'Primary')]), 100) if get_header_index(data.headers, 'Primary') != -1 else None,
                                    'middle_school': min(float(row_data[get_header_index(data.headers, 'Middle School')]), 100) if get_header_index(data.headers, 'Middle School') != -1 else None,
                                    'high_school': min(float(row_data[get_header_index(data.headers, 'High School')]), 100) if get_header_index(data.headers, 'High School') != -1 else None,
                                    'university': min(float(row_data[get_header_index(data.headers, 'University')]), 100) if get_header_index(data.headers, 'University') != -1 else None,
                                }
                            
                            # Vérifier si au moins une donnée d'éducation est présente
                            if any(education_row.get(field) is not None for field in ['no_education', 'preschool', 'primary', 'middle_school', 'high_school', 'university']):
                                education_data_rows.append(education_row)
                                education_imported += 1
                            else:
                                education_skipped += 1
                                
                        except (ValueError, TypeError) as e:
                            education_skipped += 1
                            continue

                # Import des données d'éducation
                if education_data_rows:
                    education_dataset = Dataset()
                    education_dataset.dict = education_data_rows
                    education_result = education_resource.import_data(education_dataset, dry_run=False, raise_errors=True)
                    messages.success(self.request, f"Les données ont été importées avec succès ({education_imported} lignes d'éducation importées)")
                else:
                    messages.warning(self.request, "Aucune donnée d'éducation n'a pu être importée")
            else:
                messages.warning(self.request, "Aucune donnée n'a pu être importée")

            # Message de fin d'import
            messages.success(self.request, "Importation terminée !")
            return super().form_valid(form)

        except Exception as e:
            messages.error(self.request, f"Une erreur est survenue lors de l'importation: {str(e)}")
            return self.render_to_response(self.get_context_data(form=form))


class CountryViewSet(viewsets.ModelViewSet):
    queryset = Country.objects.all()
    serializer_class = CountrySerializer

class RegionViewSet(viewsets.ModelViewSet):
    queryset = Region.objects.all()
    serializer_class = RegionSerializer

class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer

class CommuneViewSet(viewsets.ModelViewSet):
    queryset = Commune.objects.all()
    serializer_class = CommuneSerializer

class DemographicDataListCreateView(generics.ListCreateAPIView):
    serializer_class = DemographicDataSerializer

    def get_queryset(self):
        queryset = DemographicData.objects.all()
        year = self.request.query_params.get('year', None)
        
        if year is not None:
            try:
                year = int(year)
                queryset = queryset.filter(census__year=year)
            except ValueError:
                # Si l'année n'est pas un nombre valide, retourner un queryset vide
                return DemographicData.objects.none()
        else:
            # Par défaut, retourner les données du recensement le plus récent
            latest_census = Census.objects.order_by('-year').first()
            if latest_census:
                queryset = queryset.filter(census=latest_census)
        
        return queryset

class DemographicDataDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = DemographicData.objects.all()
    serializer_class = DemographicDataSerializer

@staff_member_required
def download_template(request):
    """Télécharger le modèle Excel pour l'import des données (avec niveau_donnee explicite)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Données à Remplir"

    # En-têtes avec descriptions (reprend le template d'origine)
    headers = [
        ('Country Code', 'Code du pays (obligatoire)'),
        ('Country Name', 'Nom du pays (obligatoire)'),
        ('Region Code', 'Code de la région (obligatoire)'),
        ('Region Name', 'Nom de la région (obligatoire)'),
        ('Department Code', 'Code du département (optionnel)'),
        ('Department Name', 'Nom du département (optionnel)'),
        ('Commune Code', 'Code de la commune (optionnel)'),
        ('Commune Name', 'Nom de la commune (optionnel)'),
        ('niveau_donnee', "Niveau de la donnée (region/departement/commune) - obligatoire"),
        # Données démographiques
        ('Total Population', 'Population totale (nombre entier)'),
        ('Male Population', 'Pourcentage de population masculine (0-100)'),
        ('Female Population', 'Pourcentage de population féminine (0-100)'),
        ('Urban Population', 'Pourcentage de population urbaine (0-100)'),
        ('Rural Population', 'Pourcentage de population rurale (0-100)'),
        ('Population 10+', 'Population de 10 ans et plus (nombre entier)'),
        ('Single Rate', 'Taux de célibataires (0-100)'),
        ('Married Rate', 'Taux de mariés (0-100)'),
        ('Divorced Rate', 'Taux de divorcés (0-100)'),
        ('Widowed Rate', 'Taux de veufs/veuves (0-100)'),
        ('School Enrollment Rate', 'Taux de scolarisation (0-100)'),
        ('Illiteracy Rate 10+', 'Taux d\'analphabétisme 10+ (0-100)'),
        ('Population 15+', 'Population de 15 ans et plus (nombre entier)'),
        ('Illiteracy Rate 15+', 'Taux d\'analphabétisme 15+ (0-100)'),
        # Données d'éducation (optionnelles)
        ('No Education', 'Pourcentage sans éducation (0-100)'),
        ('Preschool', 'Pourcentage préscolaire (0-100)'),
        ('Primary', 'Pourcentage primaire (0-100)'),
        ('Middle School', 'Pourcentage collège (0-100)'),
        ('High School', 'Pourcentage lycée (0-100)'),
        ('University', 'Pourcentage université (0-100)')
    ]

    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    description_font = Font(italic=True, color="666666")
    description_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Ajouter les en-têtes et descriptions
    for col, (header, description) in enumerate(headers, 1):
        # En-tête
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        # Description
        desc_cell = ws.cell(row=2, column=col, value=description)
        desc_cell.font = description_font
        desc_cell.alignment = description_alignment

    row = 3
    # 1. Régions
    for region in Region.objects.select_related('country').all():
        ws.cell(row=row, column=1, value=region.country.code if region.country else '')
        ws.cell(row=row, column=2, value=region.country.name if region.country else '')
        ws.cell(row=row, column=3, value=region.adm1_pcode)
        ws.cell(row=row, column=4, value=region.adm1_en)
        ws.cell(row=row, column=9, value='region')
        # Le reste vide
        for col in list(range(5, 9)) + list(range(10, 30)):
            ws.cell(row=row, column=col, value='')
        row += 1
    # 2. Départements
    for department in Department.objects.select_related('region__country').all():
        ws.cell(row=row, column=1, value=department.region.country.code if department.region and department.region.country else '')
        ws.cell(row=row, column=2, value=department.region.country.name if department.region and department.region.country else '')
        ws.cell(row=row, column=3, value=department.region.adm1_pcode if department.region else '')
        ws.cell(row=row, column=4, value=department.region.adm1_en if department.region else '')
        ws.cell(row=row, column=5, value=department.adm2_pcode)
        ws.cell(row=row, column=6, value=department.adm2_en)
        ws.cell(row=row, column=9, value='departement')
        # Le reste vide
        for col in [7,8] + list(range(10, 30)):
            ws.cell(row=row, column=col, value='')
        row += 1
    # 3. Communes
    for commune in Commune.objects.select_related('department__region__country').all():
        ws.cell(row=row, column=1, value=commune.department.region.country.code if commune.department and commune.department.region and commune.department.region.country else '')
        ws.cell(row=row, column=2, value=commune.department.region.country.name if commune.department and commune.department.region and commune.department.region.country else '')
        ws.cell(row=row, column=3, value=commune.department.region.adm1_pcode if commune.department and commune.department.region else '')
        ws.cell(row=row, column=4, value=commune.department.region.adm1_en if commune.department and commune.department.region else '')
        ws.cell(row=row, column=5, value=commune.department.adm2_pcode if commune.department else '')
        ws.cell(row=row, column=6, value=commune.department.adm2_en if commune.department else '')
        ws.cell(row=row, column=7, value=commune.adm3_pcode)
        ws.cell(row=row, column=8, value=commune.adm3_en)
        ws.cell(row=row, column=9, value='commune')
        # Le reste vide
        for col in range(10, 30):
            ws.cell(row=row, column=col, value='')
        row += 1

    # Ajuster la largeur des colonnes
    for col in range(1, 30):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Instructions
    instruction_row = row + 2
    ws.cell(row=instruction_row, column=1, value="INSTRUCTIONS :")
    ws.cell(row=instruction_row + 1, column=1, value="- Remplir les colonnes de données pour chaque ligne selon le niveau indiqué dans 'niveau_donnee'.")
    ws.cell(row=instruction_row + 2, column=1, value="- Ne pas modifier les colonnes administratives ni 'niveau_donnee'.")
    ws.cell(row=instruction_row + 3, column=1, value="- Vous pouvez ajouter d'autres lignes si besoin, mais respectez les valeurs possibles pour 'niveau_donnee' : region, departement, commune.")

    instruction_font = Font(bold=True, color="000000")
    for r in range(instruction_row, instruction_row + 4):
        cell = ws.cell(row=r, column=1)
        cell.font = instruction_font

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=template_import_donnees_niveaux.xlsx'
    wb.save(response)
    return response

def export_all_data(request):
    year = request.GET.get('year')
    if year:
        try:
            census = Census.objects.get(year=year)
        except Census.DoesNotExist:
            census = None
    else:
        census = Census.objects.order_by('-year').first()
        year = census.year if census else ''

    wb = openpyxl.Workbook()

    # Feuille pour les régions
    ws_region = wb.active
    ws_region.title = "Régions"
    ws_region.append([
        'Nom Région', 'Code Région',
        'Population Totale', 'Pourcentage Hommes', 'Pourcentage Femmes',
        'Population 10+', 'Taux Célibataire', 'Taux Marié',
        'Taux Divorcé', 'Taux Veuf', "Taux Scolarisation",
        "Taux d'Analphabétisme (10+)", "Population 15+",
        "Taux d'Analphabétisme (15+)",
        'Aucun niveau', 'Préscolaire', 'Primaire', 'Collège', 'Lycée', 'Université'
    ])
    
    for cell in ws_region[1]:
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    regions = Region.objects.all()
    for region in regions:
        demo_data = DemographicData.objects.filter(region=region, census=census).first() if census else None
        if demo_data:
            education = getattr(demo_data, 'educationlevel', None)
            ws_region.append([
                region.adm1_en, region.adm1_pcode,
                demo_data.total_population, demo_data.male_percentage, demo_data.female_percentage,
                demo_data.population_10_plus, demo_data.single_rate, demo_data.married_rate,
                demo_data.divorced_rate, demo_data.widowed_rate, demo_data.school_enrollment_rate,
                demo_data.illiteracy_rate_10_plus, demo_data.population_15_plus,
                demo_data.illiteracy_rate_15_plus,
                education.no_education if education else None,
                education.preschool if education else None,
                education.primary if education else None,
                education.middle_school if education else None,
                education.high_school if education else None,
                education.university if education else None,
            ])
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws_region.column_dimensions[col].width = 20

    ws_department = wb.create_sheet(title="Départements")
    ws_department.append([
        'Région', 'Nom Département', 'Code Département',
        'Population Totale', 'Pourcentage Hommes', 'Pourcentage Femmes',
        'Population 10+', 'Taux Célibataire', 'Taux Marié',
        'Taux Divorcé', 'Taux Veuf', "Taux Scolarisation",
        "Taux d'Analphabétisme (10+)", "Population 15+",
        "Taux d'Analphabétisme (15+)",
        'Aucun niveau', 'Préscolaire', 'Primaire', 'Collège', 'Lycée', 'Université'
    ])
    for cell in ws_department[1]:
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    departments = Department.objects.all()
    for department in departments:
        demo_data = DemographicData.objects.filter(department=department, census=census).first() if census else None
        if demo_data:
            education = getattr(demo_data, 'educationlevel', None)
            ws_department.append([
                department.region.adm1_en, department.adm2_en, department.adm2_pcode,
                demo_data.total_population, demo_data.male_percentage, demo_data.female_percentage,
                demo_data.population_10_plus, demo_data.single_rate, demo_data.married_rate,
                demo_data.divorced_rate, demo_data.widowed_rate, demo_data.school_enrollment_rate,
                demo_data.illiteracy_rate_10_plus, demo_data.population_15_plus,
                demo_data.illiteracy_rate_15_plus,
                education.no_education if education else None,
                education.preschool if education else None,
                education.primary if education else None,
                education.middle_school if education else None,
                education.high_school if education else None,
                education.university if education else None,
            ])
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws_department.column_dimensions[col].width = 20

    ws_commune = wb.create_sheet(title="Communes")
    ws_commune.append([
        'Département', 'Nom Commune', 'Code Commune',
        'Population Totale', 'Pourcentage Hommes', 'Pourcentage Femmes',
        'Population 10+', 'Taux Célibataire', 'Taux Marié',
        'Taux Divorcé', 'Taux Veuf', "Taux Scolarisation",
        "Taux d'Analphabétisme (10+)", "Population 15+",
        "Taux d'Analphabétisme (15+)",
        'Aucun niveau', 'Préscolaire', 'Primaire', 'Collège', 'Lycée', 'Université'
    ])
    for cell in ws_commune[1]:
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    communes = Commune.objects.all()
    for commune in communes:
        demo_data = DemographicData.objects.filter(commune=commune, census=census).first() if census else None
        if demo_data:
            education = getattr(demo_data, 'educationlevel', None)
            ws_commune.append([
                commune.department.adm2_en, commune.adm3_en, commune.adm3_pcode,
                demo_data.total_population, demo_data.male_percentage, demo_data.female_percentage,
                demo_data.population_10_plus, demo_data.single_rate, demo_data.married_rate,
                demo_data.divorced_rate, demo_data.widowed_rate, demo_data.school_enrollment_rate,
                demo_data.illiteracy_rate_10_plus, demo_data.population_15_plus,
                demo_data.illiteracy_rate_15_plus,
                education.no_education if education else None,
                education.preschool if education else None,
                education.primary if education else None,
                education.middle_school if education else None,
                education.high_school if education else None,
                education.university if education else None,
            ])
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws_commune.column_dimensions[col].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="donnees_mauritanie_{year}.xlsx"'
    wb.save(response)
    return response

def export_zone_data(request, zone_id, zone_type):
    year = request.GET.get('year')
    if year:
        try:
            census = Census.objects.get(year=year)
        except Census.DoesNotExist:
            census = None
    else:
        census = Census.objects.order_by('-year').first()
        year = census.year if census else ''

    wb = Workbook()
    ws = wb.active
    
    headers = [
        'Zone',
        'Type',
        'Population Totale',
        'Pourcentage Hommes',
        'Pourcentage Femmes',
        'Population 10+',
        'Taux Célibataire',
        'Taux Marié',
        'Taux Divorcé',
        'Taux Veuf',
        "Taux Scolarisation",
        "Taux d'Analphabétisme (10+)",
        "Population 15+",
        "Taux d'Analphabétisme (15+)"
    ]
    
    zone = None
    zone_name = ""
    data = None
    
    if zone_type == 'region':
        zone = Region.objects.get(id=zone_id)
        zone_name = zone.adm1_en
        data = DemographicData.objects.filter(region=zone, census=census).first() if census else None
        ws.title = f"Région {zone_name}"
        
    elif zone_type == 'department':
        zone = Department.objects.get(id=zone_id)
        zone_name = zone.adm2_en
        data = DemographicData.objects.filter(department=zone, census=census).first() if census else None
        ws.title = f"Département {zone_name}"
        
    elif zone_type == 'commune':
        zone = Commune.objects.get(id=zone_id)
        zone_name = zone.adm3_en
        data = DemographicData.objects.filter(commune=zone, census=census).first() if census else None
        ws.title = f"Commune {zone_name}"
        
    else:
        return HttpResponse("Type de zone non valide.", status=400)

    ws.append(headers)
    
    if data:
        ws.append([
            zone_name,
            zone_type.capitalize(),
            data.total_population,
            data.male_percentage,
            data.female_percentage,
            data.population_10_plus,
            data.single_rate,
            data.married_rate,
            data.divorced_rate,
            data.widowed_rate,
            data.school_enrollment_rate,
            data.illiteracy_rate_10_plus,
            data.population_15_plus,
            data.illiteracy_rate_15_plus
        ])
    else:
        ws.append([zone_name, zone_type.capitalize()] + ['N/A'] * (len(headers) - 2))

    for cell in ws[1]:
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    if ws.max_row > 1:
        for cell in ws[2]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    column_widths = {
        'A': 25,  # Nom de la zone
        'B': 15,  # Type
        'C': 20, 'D': 20, 'E': 20, 'F': 15, 
        'G': 15, 'H': 15, 'I': 15, 'J': 15,
        'K': 20, 'L': 25, 'M': 15, 'N': 25
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    filename = f"donnees_{zone_type}_{zone_name.replace(' ', '_')}_{year}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

class CensusYearsView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        available_years = list(Census.objects.values_list('year', flat=True))
        return Response(available_years)